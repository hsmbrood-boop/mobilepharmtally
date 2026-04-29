"""Focused inspection of rows 12..27 across multiple sample xlsx files."""
import sys
from openpyxl import load_workbook

paths = sys.argv[1:]
for path in paths:
    wb = load_workbook(path, data_only=False)
    ws = wb['정산내역']
    print(f"\n=== {path} (max_row={ws.max_row}) ===")
    print(f"merged_ranges: {[str(r) for r in ws.merged_cells.ranges]}")
    print("row | A | B | C | D")
    for row_idx in range(1, ws.max_row + 1):
        a = ws.cell(row_idx, 1).value
        b = ws.cell(row_idx, 2).value
        c = ws.cell(row_idx, 3).value
        d = ws.cell(row_idx, 4).value
        if a is None and b is None and c is None and d is None:
            print(f"{row_idx:3d} | (empty)")
        else:
            print(f"{row_idx:3d} | {a!r:30} | {b!r:14} | {c!r:30} | {d!r}")
