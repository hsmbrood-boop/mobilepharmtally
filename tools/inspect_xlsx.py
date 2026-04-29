"""Inspect a PharmTally sample xlsx so we can replicate the format exactly."""
import sys
from openpyxl import load_workbook

path = sys.argv[1]
wb = load_workbook(path, data_only=False)

print(f"=== FILE: {path} ===")
print(f"Sheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    print(f"dimensions: {ws.dimensions}")
    print(f"max_row: {ws.max_row}, max_col: {ws.max_column}")
    print(f"merged_ranges: {[str(r) for r in ws.merged_cells.ranges]}")

    print("\n[Cells: row,col -> value | number_format | font | fill | alignment | border]")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                             min_col=1, max_col=ws.max_column):
        for c in row:
            if c.value is None and c.number_format == 'General' and not c.has_style:
                continue
            font = f"{c.font.name}/{c.font.size}/bold={c.font.bold}/color={c.font.color.rgb if c.font.color else None}"
            fill = c.fill.fgColor.rgb if c.fill and c.fill.fgColor else None
            align = f"h={c.alignment.horizontal} v={c.alignment.vertical} wrap={c.alignment.wrap_text}"
            print(f"  {c.coordinate}: value={c.value!r:40} fmt={c.number_format!r:18} font={font} fill={fill} align={align}")

    print("\n[Column widths]")
    for col_letter, col_dim in ws.column_dimensions.items():
        print(f"  {col_letter}: width={col_dim.width}")
    print("\n[Row heights]")
    for row_idx, row_dim in ws.row_dimensions.items():
        print(f"  row{row_idx}: height={row_dim.height}")
